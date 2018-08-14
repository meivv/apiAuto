#-*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
class api_rd_wrt_excel(object):
    def __init__(self,excl_path,model='r'):
        try:
            self.path = excl_path
            # self.sheet_name = sht_name
            if model=='r':
                self.excel=openpyxl.load_workbook(self.path)
            elif model=='w':
                self.excel=openpyxl.Workbook()
        except Exception as e:
            print(e)

    #获取EXCEL的sheet
    def excel_get_sheet(self,sht_name):
        return  self.excel[sht_name]

    #创建新sheet
    def excel_create_sheet(self,sht_name):
        return  self.excel.create_sheet(sht_name)

    #获取EXCEL的总行数
    def excel_get_rownum(self,sheet):
        try:
            return sheet.max_row
        except Exception as e:
            print(e)
            return None

    #获取EXCEL的总列数
    def excel_get_colnum(self,sheet):
        try:
            return sheet.max_column
        except Exception as e:
            print(e)
            return None

    #获取EXCEL某个单元格的值
    def excel_get_cell_value(self,sheet,r,c):
        try:
            return sheet.cell(row=r,column=c).value
        except Exception as e:
            print(e)
            return None

    #获取EXCEL某一行的值
    def excel_get_row_value(self,sheet,r):
        try:
            lst=[]
            for cell in list(sheet.rows)[r]:
                lst.append(cell.value)
            return lst
        except Exception as e:
            print(e)
            return None

    #获取EXCEL某一列的值
    def excel_get_col_value(self,sheet,c):
        try:
            lst=[]
            for col in list(sheet.columns)[c]:
                lst.append(col.value)
            return lst
        except Exception as e:
            print(e)
            return None

    #EXCEL设置某一单元格值
    def excel_wrt_cell_value(self,sheet,r,c,v):
        try:
            sheet.cell(row=r,column=c,value=v)
        except Exception as e:
            print(e)

    #EXCEL设置某一行的值
    def excel_wrt_row_value(self,sheet,row_value):
        try:
            sheet.append(row_value)
        except Exception as e:
            print(e)

    #保存EXCEL
    def excel_save(self):
        try:
            self.excel.save(self.path)
        except Exception as e:
            print(e)

